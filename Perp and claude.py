
import os

code = r'''#!/usr/bin/env python3
"""
NSE Options Live Dashboard v3.0
Powered by Dhan API v2
"""

# ── IMPORTS ──────────────────────────────────────────────────
import requests
import openpyxl
import openpyxl.styles as _oxl_styles
import openpyxl.utils as _oxl_utils
import datetime, os, time, csv, signal

from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.columns import Columns
from rich import box

try:
    import winsound
    _WINSOUND_OK = True
except Exception:
    _WINSOUND_OK = False
    winsound = None

try:
    from plyer import notification as _plyer_notify
    _PLYER_OK = True
except Exception:
    _PLYER_OK = False
    _plyer_notify = None

# ── SETTINGS (edit these) ─────────────────────────────────────
CLIENT_ID     = "1110828392"
ACCESS_TOKEN  = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzUxMiJ9.eyJpc3MiOiJkaGFuIiwicGFydG5lcklkIjoiIiwiZXhwIjoxNzc1NDkwOTg2LCJpYXQiOjE3NzU0MDQ1ODYsInRva2VuQ29uc3VtZXJUeXBlIjoiU0VMRiIsIndlYmhvb2tVcmwiOiIiLCJkaGFuQ2xpZW50SWQiOiIxMTEwODI4MzkyIn0.va3GwEoTPudAXGGEMN7HPWGZP9sfv3QiNpReCVeEJkoVBInYIYqmfKFJPjPjQax2F36r3c5WMuUfyn2LR_ptWA"

SYMBOL        = "NIFTY"       # NIFTY | BANKNIFTY | FINNIFTY | RELIANCE | SBIN
EXPIRY_DATE   = None          # None = auto nearest | or "10-Apr-2026"
NUM_STRIKES   = 8             # strikes above & below ATM
REFRESH_EVERY = 5             # seconds between refreshes (min 3)

SAVE_TO_CSV           = True
DESKTOP_NOTIFICATIONS = True
ALERT_PCR_BEARISH     = 1.3
ALERT_PCR_BULLISH     = 0.7
ALERT_OI_SPIKE_PCT    = 20

WATCHLIST_MODE = False
WATCHLIST      = ["NIFTY", "BANKNIFTY", "RELIANCE"]

# ── DHAN API ──────────────────────────────────────────────────
BASE_URL = "https://api.dhan.co/v2"

SYMBOL_MAP = {
    "NIFTY":     {"scrip_code": 13,   "segment": "IDX_I"},
    "BANKNIFTY": {"scrip_code": 25,   "segment": "IDX_I"},
    "FINNIFTY":  {"scrip_code": 27,   "segment": "IDX_I"},
    "RELIANCE":  {"scrip_code": 1333, "segment": "NSE_EQ"},
    "SBIN":      {"scrip_code": 3045, "segment": "NSE_EQ"},
}

HEADERS = {
    "access-token": ACCESS_TOKEN,
    "client-id": CLIENT_ID,
    "Content-Type": "application/json",
}

# ── GLOBAL STATE ──────────────────────────────────────────────
class State:
    def __init__(self):
        self.prev_rows      = {}
        self.pcr_history    = {}
        self.recent_alerts  = []
        self.all_snapshots  = []
        self.csv_file       = None
        self.csv_written    = False
        self.shutdown_req   = False
        self.api_tested     = False
        self.last_data_ts   = None

_state  = State()
console = Console()

# ── HTTP HELPER ───────────────────────────────────────────────
def make_request(method, endpoint, body=None):
    url = BASE_URL + endpoint
    try:
        if method.upper() == "POST":
            resp = requests.post(url, json=body, headers=HEADERS, timeout=30)
        else:
            resp = requests.get(url, headers=HEADERS, timeout=30)

        if resp.status_code == 401:
            console.print("[red][ERROR] 401 Unauthorized. Check CLIENT_ID / ACCESS_TOKEN.[/red]")
            raise SystemExit(1)

        if resp.status_code == 429:
            console.print("[yellow][WARNING] Rate limited (429). Waiting 10s...[/yellow]")
            time.sleep(10)
            resp = requests.post(url, json=body, headers=HEADERS, timeout=30)

        if resp.status_code in (500, 502, 503):
            console.print(f"[yellow][WARNING] Server error {resp.status_code}. Waiting 30s...[/yellow]")
            time.sleep(30)
            resp = requests.post(url, json=body, headers=HEADERS, timeout=30)

        if resp.status_code != 200:
            console.print(f"[red][ERROR] API returned {resp.status_code}: {resp.text}[/red]")
            raise SystemExit(1)

        return resp.json()

    except requests.exceptions.ConnectionError:
        console.print("[red][ERROR] No internet connection.[/red]")
        raise SystemExit(1)
    except requests.exceptions.Timeout:
        console.print("[red][ERROR] Request timed out.[/red]")
        raise SystemExit(1)

# ── API FUNCTIONS ─────────────────────────────────────────────
def get_expiry_list(symbol):
    if symbol not in SYMBOL_MAP:
        console.print(f"[red]Unknown symbol: {symbol}[/red]")
        raise SystemExit(1)
    body = {
        "UnderlyingScrip": SYMBOL_MAP[symbol]["scrip_code"],
        "UnderlyingSeg": SYMBOL_MAP[symbol]["segment"],
    }
    data = make_request("POST", "/optionchain/expirylist", body)
    if "data" in data and isinstance(data["data"], list):
        return data["data"]
    console.print(f"[red]Unexpected expiry response: {data}[/red]")
    return []

def get_option_chain(symbol, expiry):
    if symbol not in SYMBOL_MAP:
        console.print(f"[red]Unknown symbol: {symbol}[/red]")
        raise SystemExit(1)
    body = {
        "UnderlyingScrip": SYMBOL_MAP[symbol]["scrip_code"],
        "UnderlyingSeg": SYMBOL_MAP[symbol]["segment"],
        "Expiry": expiry,
    }
    return make_request("POST", "/optionchain", body)

# ── CALCULATIONS ──────────────────────────────────────────────
def get_atm_strike(spot_price, strikes):
    if not strikes:
        return 0
    return min(strikes, key=lambda x: abs(x - spot_price))

def calc_pcr(rows):
    total_ce = sum(r["ce_oi"] for r in rows)
    total_pe = sum(r["pe_oi"] for r in rows)
    if total_ce == 0:
        return 0.0
    return round(total_pe / total_ce, 3)

def calc_max_pain(rows):
    if not rows:
        return 0
    row_by_strike = {r["strike"]: r for r in rows}
    strikes = sorted(row_by_strike.keys())
    pain_by_strike = {}
    for i, test_k in enumerate(strikes):
        pain = 0.0
        for k in strikes[:i]:
            pain += row_by_strike[k]["ce_oi"] * max(0.0, test_k - k)
        for k in strikes[i + 1:]:
            pain += row_by_strike[k]["pe_oi"] * max(0.0, k - test_k)
        pain_by_strike[test_k] = pain
    return min(pain_by_strike, key=lambda k: pain_by_strike[k])

def get_oi_change(symbol, rows):
    prev_rows = _state.prev_rows.get(symbol, [])
    prev_lookup = {r["strike"]: r for r in prev_rows} if prev_rows else {}
    changes = {}
    for row in rows:
        strike = row["strike"]
        prev = prev_lookup.get(strike, {})
        prev_ce = prev.get("ce_oi", 0)
        prev_pe = prev.get("pe_oi", 0)
        ce_abs = row["ce_oi"] - prev_ce
        pe_abs = row["pe_oi"] - prev_pe
        ce_pct = _pct_change(row["ce_oi"], prev_ce)
        pe_pct = _pct_change(row["pe_oi"], prev_pe)
        changes[strike] = {
            "ce_abs": ce_abs, "pe_abs": pe_abs,
            "ce_pct": ce_pct, "pe_pct": pe_pct,
        }
    _state.prev_rows[symbol] = list(rows)
    return changes

def _pct_change(current, previous):
    if previous == 0:
        return 0.0 if current == 0 else 100.0
    return round(((current - previous) / previous) * 100, 1)

# ── PARSE OPTION CHAIN ────────────────────────────────────────
def parse_option_chain(raw_data, symbol):
    try:
        data = raw_data.get("data", {})
        spot_price = data.get("last_price", 0)
        oc = data.get("oc", {})
        rows = []
        for strike_str, ce_pe in oc.items():
            strike = float(strike_str)
            ce = ce_pe.get("ce", {}) or ce_pe.get("CE", {})
            ce_g = ce.get("greeks", {}) or {}
            pe = ce_pe.get("pe", {}) or ce_pe.get("PE", {})
            pe_g = pe.get("greeks", {}) or {}
            rows.append({
                "strike":    strike,
                "symbol":    symbol,
                "ce_oi":     ce.get("oi", 0),
                "ce_ltp":    ce.get("last_price", 0),
                "ce_iv":     ce.get("implied_volatility", 0) or ce.get("iv", 0),
                "ce_volume": ce.get("volume", 0),
                "ce_bid":    ce.get("top_bid_price", 0),
                "ce_ask":    ce.get("top_ask_price", 0),
                "ce_delta":  ce_g.get("delta", 0),
                "ce_theta":  ce_g.get("theta", 0),
                "ce_gamma":  ce_g.get("gamma", 0),
                "ce_vega":   ce_g.get("vega", 0),
                "pe_oi":     pe.get("oi", 0),
                "pe_ltp":    pe.get("last_price", 0),
                "pe_iv":     pe.get("implied_volatility", 0) or pe.get("iv", 0),
                "pe_volume": pe.get("volume", 0),
                "pe_bid":    pe.get("top_bid_price", 0),
                "pe_ask":    pe.get("top_ask_price", 0),
                "pe_delta":  pe_g.get("delta", 0),
                "pe_theta":  pe_g.get("theta", 0),
                "pe_gamma":  pe_g.get("gamma", 0),
                "pe_vega":   pe_g.get("vega", 0),
            })
        rows.sort(key=lambda r: r["strike"])
        return spot_price, rows
    except Exception as e:
        console.print(f"[red]Parse error: {e}[/red]")
        return 0, []

# ── FORMATTING HELPERS ────────────────────────────────────────
def fmt_oi(value):
    if value is None:
        return "—"
    v = float(value)
    if abs(v) >= 1_000_000:
        return f"{v / 1_000_000:.2f}M"
    elif abs(v) >= 1_000:
        return f"{v / 1_000:.0f}K"
    elif v == 0:
        return "—"
    return f"{v:.0f}"

def fmt_chg_abs(value):
    if value is None or value == 0:
        return "[dim]—[/dim]"
    v = float(value)
    arrow = "▲" if v > 0 else "▼"
    colour = "green" if v > 0 else "red"
    return f"[{colour}]{arrow}{fmt_oi(abs(v))}[/{colour}]"

def fmt_float(value, decimals=2, prefix="", suffix=""):
    if value is None or value == 0:
        return "—"
    return f"{prefix}{value:.{decimals}f}{suffix}"

# ── ALERTS ────────────────────────────────────────────────────
def trigger_alert(message, beeps=3):
    for _ in range(beeps):
        if _WINSOUND_OK:
            winsound.Beep(1000, 250)
            time.sleep(0.1)
        else:
            print("\a", end="")
            time.sleep(0.3)
    if DESKTOP_NOTIFICATIONS and _PLYER_OK:
        try:
            _plyer_notify.notify(
                title="Dhan Dashboard Alert",
                message=message,
                app_name="Dhan Dashboard",
                timeout=8,
            )
        except Exception:
            pass
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    _state.recent_alerts.append(f"[{ts}] {message}")
    if len(_state.recent_alerts) > 10:
        _state.recent_alerts.pop(0)

def check_alerts(symbol, _, pcr, oi_changes):
    triggered = []
    pcr_hist = _state.pcr_history.get(symbol, [])
    prev_pcr = pcr_hist[-1] if pcr_hist else None
    pcr_hist.append(pcr)
    if len(pcr_hist) > 30:
        pcr_hist.pop(0)
    _state.pcr_history[symbol] = pcr_hist
    if prev_pcr is not None:
        if prev_pcr <= ALERT_PCR_BEARISH and pcr > ALERT_PCR_BEARISH:
            msg = f"PCR crossed {ALERT_PCR_BEARISH} → Bearish Signal!"
            trigger_alert(msg, beeps=3)
            triggered.append(msg)
        if prev_pcr >= ALERT_PCR_BULLISH and pcr < ALERT_PCR_BULLISH:
            msg = f"PCR fell to {pcr:.2f} → Bullish Signal!"
            trigger_alert(msg, beeps=3)
            triggered.append(msg)
    for strike, chg in oi_changes.items():
        for side, pct_key, abs_key in [("CE", "ce_pct", "ce_abs"), ("PE", "pe_pct", "pe_abs")]:
            pct = chg[pct_key]
            abs_ = chg[abs_key]
            if abs(pct) >= ALERT_OI_SPIKE_PCT:
                arrow = "▲" if pct > 0 else "▼"
                msg = f"{side} OI spike at ₹{strike:.0f}: {arrow}{abs(pct):.0f}% ({fmt_oi(abs_)})"
                trigger_alert(msg, beeps=1)
                triggered.append(msg)
    return triggered

# ── CSV ───────────────────────────────────────────────────────
CSV_COLUMNS = [
    "Time", "Symbol", "Expiry", "Spot", "Strike",
    "CE_OI", "CE_LTP", "CE_IV", "CE_Delta", "CE_Theta", "CE_Vega", "CE_Volume",
    "PE_OI", "PE_LTP", "PE_IV", "PE_Delta", "PE_Theta", "PE_Vega", "PE_Volume",
    "PCR", "Max_Pain",
]

def save_to_csv(timestamp, symbol, expiry, spot, rows, pcr, max_pain):
    today_str = datetime.date.today().strftime("%Y-%m-%d")
    if _state.csv_file is None:
        _state.csv_file = f"{symbol}_options_{today_str}.csv"
    file_exists = os.path.exists(_state.csv_file)
    try:
        with open(_state.csv_file, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS)
            if not file_exists:
                writer.writeheader()
            _state.csv_written = True
            for row in rows:
                writer.writerow({
                    "Time": timestamp, "Symbol": symbol, "Expiry": expiry,
                    "Spot": f"{spot:.2f}", "Strike": f"{row['strike']:.0f}",
                    "CE_OI": int(row["ce_oi"]), "CE_LTP": f"{row['ce_ltp']:.2f}",
                    "CE_IV": f"{row['ce_iv']:.2f}", "CE_Delta": f"{row['ce_delta']:.4f}",
                    "CE_Theta": f"{row['ce_theta']:.2f}", "CE_Vega": f"{row['ce_vega']:.2f}",
                    "CE_Volume": int(row["ce_volume"]),
                    "PE_OI": int(row["pe_oi"]), "PE_LTP": f"{row['pe_ltp']:.2f}",
                    "PE_IV": f"{row['pe_iv']:.2f}", "PE_Delta": f"{row['pe_delta']:.4f}",
                    "PE_Theta": f"{row['pe_theta']:.2f}", "PE_Vega": f"{row['pe_vega']:.2f}",
                    "PE_Volume": int(row["pe_volume"]),
                    "PCR": f"{pcr:.4f}", "Max_Pain": f"{max_pain:.0f}",
                })
    except Exception as e:
        console.print(f"[yellow]CSV write error: {e}[/yellow]")

# ── EXCEL EXPORT ──────────────────────────────────────────────
def export_excel(snapshots, symbol):
    ts_str   = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{symbol}_DhanReport_{ts_str}.xlsx"
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    S = _oxl_styles
    thin = S.Border(
        left=S.Side(style="thin"), right=S.Side(style="thin"),
        top=S.Side(style="thin"), bottom=S.Side(style="thin"),
    )
    hdr_fill  = S.PatternFill("solid", fgColor="1F4E79")
    hdr_font  = S.Font(bold=True, color="FFFFFF")
    hdr_align = S.Alignment(horizontal="center", vertical="center")
    ce_fill   = S.PatternFill("solid", fgColor="D6E4F0")
    pe_fill   = S.PatternFill("solid", fgColor="FADBD8")
    atm_fill  = S.PatternFill("solid", fgColor="FCF3CF")
    pain_fill = S.PatternFill("solid", fgColor="E8DAEF")

    def style_header(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = hdr_align
            cell.border = thin

    ws1 = wb.create_sheet(title="Options Chain")
    chain_cols = [
        "Strike", "CE_OI", "CE_Volume", "CE_LTP", "CE_Bid", "CE_Ask",
        "CE_IV", "CE_Delta", "CE_Theta", "CE_Gamma", "CE_Vega",
        "PE_Vega", "PE_Gamma", "PE_Theta", "PE_Delta",
        "PE_IV", "PE_Ask", "PE_Bid", "PE_LTP", "PE_Volume", "PE_OI",
    ]
    for col, h in enumerate(chain_cols, 1):
        ws1.cell(row=1, column=col, value=h)
    style_header(ws1, 1, len(chain_cols))
    row_num = 2
    for snap in snapshots:
        atm_s = snap["atm_strike"]
        mpain = snap["max_pain"]
        for r in snap["rows"]:
            strike = r["strike"]
            is_atm  = (strike == atm_s)
            is_pain = (strike == mpain)
            vals = [
                strike, r["ce_oi"], r["ce_volume"], r["ce_ltp"],
                r["ce_bid"], r["ce_ask"], r["ce_iv"], r["ce_delta"],
                r["ce_theta"], r["ce_gamma"], r["ce_vega"],
                r["pe_vega"], r["pe_gamma"], r["pe_theta"], r["pe_delta"],
                r["pe_iv"], r["pe_ask"], r["pe_bid"], r["pe_ltp"],
                r["pe_volume"], r["pe_oi"],
            ]
            for col, val in enumerate(vals, 1):
                cell = ws1.cell(row=row_num, column=col, value=val)
                cell.border = thin
                if is_pain:
                    cell.fill = pain_fill
                elif is_atm:
                    cell.fill = atm_fill
                elif col <= 11:
                    cell.fill = ce_fill
                else:
                    cell.fill = pe_fill
            row_num += 1
    ws1.column_dimensions["A"].width = 10
    for col in range(2, len(chain_cols) + 1):
        ws1.column_dimensions[_oxl_utils.get_column_letter(col)].width = 12
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet(title="PCR History")
    for col, h in enumerate(["Timestamp", "Symbol", "PCR"], 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, 1, 3)
    row_idx = 2
    for sym, hist in _state.pcr_history.items():
        for i, pcr_val in enumerate(hist):
            ws2.cell(row=row_idx, column=1, value=f"Iteration {i+1}")
            ws2.cell(row=row_idx, column=2, value=sym)
            ws2.cell(row=row_idx, column=3, value=pcr_val)
            row_idx += 1
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 10

    ws3 = wb.create_sheet(title="Alerts")
    for col, h in enumerate(["Timestamp", "Alert Message"], 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, 1, 2)
    for i, alert in enumerate(_state.recent_alerts, 2):
        ws3.cell(row=i, column=1, value=alert[:19])
        ws3.cell(row=i, column=2, value=alert[21:])
        ws3.cell(row=i, column=1).border = thin
        ws3.cell(row=i, column=2).border = thin
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 60

    try:
        wb.save(filename)
        console.print(f"\n[green]Excel saved: {filename}[/green]")
    except Exception as e:
        console.print(f"\n[red]Could not save Excel: {e}[/red]")

# ── BIAS HELPERS ──────────────────────────────────────────────
def get_bias_label(pcr):
    if pcr > 1.3:   return "STRONG BEARISH"
    elif pcr > 1.2: return "BEARISH"
    elif pcr < 0.5: return "STRONG BULLISH"
    elif pcr < 0.7: return "BULLISH"
    else:           return "NEUTRAL"

def get_bias_colour(pcr):
    if pcr < 0.7:   return "green"
    elif pcr > 1.2: return "red"
    else:           return "yellow"

# ── DASHBOARD TABLE ───────────────────────────────────────────
def build_main_table(spot, rows, atm_strike, max_pain, pcr, expiry, oi_changes):
    table = Table(
        show_header=True, header_style="bold white",
        border_style="dim blue", pad_edge=False, expand=False, box=box.ROUNDED,
    )
    table.add_column("CE OI",    justify="right", style="bold blue", min_width=9)
    table.add_column("CE Chg",   justify="right", style="blue",      min_width=8)
    table.add_column("CE Vol",   justify="right", style="blue",      min_width=7)
    table.add_column("CE LTP",   justify="right", style="blue",      min_width=8)
    table.add_column("CE IV%",   justify="right", style="blue",      min_width=7)
    table.add_column("CE Delta", justify="right", style="blue",      min_width=8)
    table.add_column("STRIKE",   justify="center", style="bold white", min_width=9)
    table.add_column("PE Delta", justify="right", style="red",       min_width=8)
    table.add_column("PE IV%",   justify="right", style="red",       min_width=7)
    table.add_column("PE LTP",   justify="right", style="red",       min_width=8)
    table.add_column("PE Vol",   justify="right", style="red",       min_width=7)
    table.add_column("PE Chg",   justify="right", style="red",       min_width=8)
    table.add_column("PE OI",    justify="right", style="bold red",  min_width=9)

    atm_index = next((i for i, r in enumerate(rows) if r["strike"] == atm_strike), None)
    if atm_index is None:
        return table

    lo = max(0, atm_index - NUM_STRIKES)
    hi = min(len(rows), atm_index + NUM_STRIKES + 1)

    for row in rows[lo:hi]:
        sv      = row["strike"]
        is_atm  = (sv == atm_strike)
        is_pain = (sv == max_pain)
        chg     = oi_changes.get(sv, {})
        if is_pain:
            marker, base = " ◆", "bold magenta"
        elif is_atm:
            marker, base = " ★", "bold yellow"
        else:
            marker, base = "", ""
        s_ce = f"{base} blue"  if base else "blue"
        s_sk = f"{base} white" if base else "white"
        s_pe = f"{base} red"   if base else "red"
        ce_abs = chg.get("ce_abs", 0) or 0
        pe_abs = chg.get("pe_abs", 0) or 0
        table.add_row(
            f"[{s_ce}]{fmt_oi(row['ce_oi'])}{marker}[/{s_ce}]",
            fmt_chg_abs(ce_abs),
            f"[{s_ce}]{fmt_oi(row['ce_volume'])}[/{s_ce}]",
            f"[{s_ce}]{fmt_float(row['ce_ltp'], 2)}[/{s_ce}]",
            f"[{s_ce}]{fmt_float(row['ce_iv'], 2)}[/{s_ce}]",
            f"[{s_ce}]{fmt_float(row['ce_delta'], 4)}[/{s_ce}]",
            f"[{s_sk}]{sv:.0f}[/{s_sk}]",
            f"[{s_pe}]{fmt_float(row['pe_delta'], 4)}[/{s_pe}]",
            f"[{s_pe}]{fmt_float(row['pe_iv'], 2)}[/{s_pe}]",
            f"[{s_pe}]{fmt_float(row['pe_ltp'], 2)}[/{s_pe}]",
            f"[{s_pe}]{fmt_oi(row['pe_volume'])}[/{s_pe}]",
            fmt_chg_abs(pe_abs),
            f"[{s_pe}]{fmt_oi(row['pe_oi'])}[/{s_pe}]",
        )
    return table

# ── PCR PANEL ─────────────────────────────────────────────────
def build_pcr_panel(pcr, pcr_hist):
    bias     = get_bias_label(pcr)
    bias_col = get_bias_colour(pcr)
    if len(pcr_hist) >= 3:
        diff = sum(pcr_hist[-2:]) / 2 - sum(pcr_hist[-3:-1]) / 2
        if diff > 0.05:
            trend_str, trend_col = "Bearish Pressure Building ▲", "red"
        elif diff < -0.05:
            trend_str, trend_col = "Bullish Pressure Building ▼", "green"
        else:
            trend_str, trend_col = "Sideways — No Clear Trend", "yellow"
    else:
        trend_str, trend_col = "Collecting data...", "dim"
    hi_str = f"[red]{max(pcr_hist):.3f}[/red]"   if pcr_hist else "[dim]—[/dim]"
    lo_str = f"[green]{min(pcr_hist):.3f}[/green]" if pcr_hist else "[dim]—[/dim]"
    history = pcr_hist[-5:] if pcr_hist else []
    arrows  = []
    if len(history) >= 2:
        for i in range(len(history) - 1):
            if   history[i+1] > history[i] + 0.01: arrows.append("[red]▲[/red]")
            elif history[i+1] < history[i] - 0.01: arrows.append("[green]▼[/green]")
            else:                                   arrows.append("[dim]—[/dim]")
    arrow_str  = " ".join(arrows) or "[dim]—[/dim]"
    hist_vals  = " ".join(f"[blue]{v:.2f}[/blue]" for v in history) or "[dim]—[/dim]"
    body = (
        f"[bold]PCR:[/bold] [bold {bias_col}]{pcr:.3f}[/bold {bias_col}]\n"
        f"[bold]Bias:[/bold] [bold {bias_col}]{bias}[/bold {bias_col}]\n\n"
        f"[bold]Last values:[/bold]\n{hist_vals}\n{arrow_str}\n\n"
        f"[bold]Trend:[/bold] [{trend_col}]{trend_str}[/{trend_col}]\n"
        f"[bold]Session Hi:[/bold] {hi_str} [bold]Lo:[/bold] {lo_str}"
    )
    return Panel(body, title="[bold]PCR Analysis[/bold]", border_style="bright_blue", padding=(1, 2))

# ── STRADDLE PANEL ────────────────────────────────────────────
def build_straddle_panel(rows, atm_strike, spot):
    atm_row = next((r for r in rows if r["strike"] == atm_strike), None)
    strikes_sorted = sorted(r["strike"] for r in rows)
    atm_idx = strikes_sorted.index(atm_strike) if atm_strike in strikes_sorted else -1
    upper_strike = strikes_sorted[atm_idx + 1] if atm_idx + 1 < len(strikes_sorted) else None
    lower_strike = strikes_sorted[atm_idx - 1] if atm_idx > 0 else None
    upper_row = next((r for r in rows if r["strike"] == upper_strike), None) if upper_strike else None
    lower_row = next((r for r in rows if r["strike"] == lower_strike), None) if lower_strike else None
    atm_ce   = atm_row["ce_ltp"] if atm_row else 0
    atm_pe   = atm_row["pe_ltp"] if atm_row else 0
    stradd   = atm_ce + atm_pe
    upper_be = atm_strike + stradd
    lower_be = atm_strike - stradd
    avg_iv   = ((atm_row["ce_iv"] if atm_row else 0) + (atm_row["pe_iv"] if atm_row else 0)) / 2
    otm_ce   = upper_row["ce_ltp"] if upper_row else 0
    otm_pe   = lower_row["pe_ltp"] if lower_row else 0
    strangle = otm_ce + otm_pe
    if avg_iv > 30:
        iv_txt, iv_note = "[red]Sell Straddle[/red]", "IV is high — premium rich"
    elif avg_iv < 20:
        iv_txt, iv_note = "[green]Buy Straddle[/green]", "IV is low — premium cheap"
    else:
        iv_txt, iv_note = "[yellow]Normal[/yellow]", "IV is fair"
    p = lambda val: f"\u20b9{val:.2f}"
    u_str = f"₹{upper_strike:.0f}" if upper_strike else "—"
    l_str = f"₹{lower_strike:.0f}" if lower_strike else "—"
    body = (
        f"[bold white]━━ ATM Straddle ━━[/bold white]\n"
        f" CE LTP : {p(atm_ce)}  PE LTP : {p(atm_pe)}\n"
        f" Straddle : [bold cyan]{p(stradd)}[/bold cyan]\n"
        f" Upper B/E: [yellow]{upper_be:.2f}[/yellow]\n"
        f" Lower B/E: [yellow]{lower_be:.2f}[/yellow]\n\n"
        f"[bold white]━━ OTM Strangle ━━[/bold white]\n"
        f" +1 CE : {u_str} → {p(otm_ce)}\n"
        f" -1 PE : {l_str} → {p(otm_pe)}\n"
        f" Strangle : [bold cyan]{p(strangle)}[/bold cyan]\n\n"
        f"[bold]ATM Spot:[/bold] {spot:.2f}  [bold]Strike:[/bold] {atm_strike:.0f}\n"
        f"[bold]Avg IV :[/bold] {avg_iv:.1f}% → {iv_txt}\n"
        f"[dim]{iv_note}[/dim]"
    )
    return Panel(body, title="[bold]Straddle / Strangle[/bold]", border_style="bright_red", padding=(1, 2))

# ── ALERTS PANEL ──────────────────────────────────────────────
def build_alerts_panel():
    alerts = _state.recent_alerts
    if not alerts:
        body = "[dim]No alerts this session[/dim]"
    else:
        shown = alerts[-5:]
        lines = []
        for i, a in enumerate(shown, 1):
            coloured = (
                a.replace("Bearish", "[red]Bearish[/red]")
                 .replace("Bullish", "[green]Bullish[/green]")
                 .replace("PCR",     "[yellow]PCR[/yellow]")
                 .replace("spike",   "[magenta]spike[/magenta]")
            )
            lines.append(f" {i}. {coloured}")
        while len(lines) < 5:
            lines.append(" [dim]—[/dim]")
        body = "\n".join(lines)
    return Panel(body, title="[bold]Recent Alerts[/bold]", border_style="bright_yellow", padding=(1, 2))

# ── WATCHLIST MODE ────────────────────────────────────────────
def run_watchlist():
    console.clear()
    console.print("\n[cyan]Loading watchlist mode...[/cyan]\n")
    symbol_expiries = {}
    for sym in WATCHLIST:
        exps = get_expiry_list(sym)
        if exps:
            symbol_expiries[sym] = sorted(exps, key=lambda d: datetime.datetime.strptime(d, "%d-%b-%Y"))[0]
        else:
            symbol_expiries[sym] = None
    iteration = 0
    while not _state.shutdown_req:
        iteration += 1
        ts = datetime.datetime.now().strftime("%H:%M:%S")
        console.clear()
        table = Table(
            show_header=True, header_style="bold white",
            border_style="dim cyan", box=box.ROUNDED,
        )
        table.add_column("Symbol",   justify="left",   style="bold white", min_width=10)
        table.add_column("Spot",     justify="right",  style="white",      min_width=9)
        table.add_column("ATM",      justify="right",  style="yellow",     min_width=8)
        table.add_column("PCR",      justify="right",  style="white",      min_width=6)
        table.add_column("Bias",     justify="center", style="white",      min_width=14)
        table.add_column("Max Pain", justify="right",  style="magenta",    min_width=9)
        table.add_column("CE OI",    justify="right",  style="blue",       min_width=8)
        table.add_column("PE OI",    justify="right",  style="red",        min_width=8)
        table.add_column("Alert",    justify="center", style="yellow",     min_width=8)
        console.print(f"[bold cyan]NSE WATCHLIST — {ts} — Refresh #{iteration}[/bold cyan]\n")
        for sym in WATCHLIST:
            expiry = symbol_expiries.get(sym)
            if not expiry:
                table.add_row(sym, "[dim]No expiry[/dim]", *["[dim]—[/dim]"] * 7)
                continue
            raw = get_option_chain(sym, expiry)
            spot, rows = parse_option_chain(raw, sym)
            if not rows:
                table.add_row(sym, "[dim]No data[/dim]", *["[dim]—[/dim]"] * 7)
                continue
            atm_strike = get_atm_strike(spot, [r["strike"] for r in rows])
            pcr   = calc_pcr(rows)
            mpain = calc_max_pain(rows)
            bias      = get_bias_label(pcr)
            bias_col  = get_bias_colour(pcr)
            total_ce  = fmt_oi(sum(r["ce_oi"] for r in rows))
            total_pe  = fmt_oi(sum(r["pe_oi"] for r in rows))
            if pcr > ALERT_PCR_BEARISH:
                alert_flag = "[red]⚠ PCR[/red]"
            elif pcr < ALERT_PCR_BULLISH:
                alert_flag = "[green]⚠ PCR[/green]"
            else:
                alert_flag = "[dim]✓[/dim]"
            if pcr > 1.2:
                sym_cell = f"[bold red]{sym}[/bold red]"
            elif pcr < 0.8:
                sym_cell = f"[bold green]{sym}[/bold green]"
            else:
                sym_cell = sym
            table.add_row(
                sym_cell, f"{spot:.2f}", f"{atm_strike:.0f}", f"{pcr:.3f}",
                f"[{bias_col}]{bias}[/{bias_col}]", f"{mpain:.0f}",
                total_ce, total_pe, alert_flag,
            )
            if sym != WATCHLIST[-1]:
                time.sleep(4)
        console.print(table)
        console.print("\n[dim]Next refresh in 90s | Press Ctrl+C to exit[/dim]")
        for _ in range(90):
            if _state.shutdown_req:
                break
            time.sleep(1)

# ── MARKET HOURS ──────────────────────────────────────────────
def is_market_open():
    now = datetime.datetime.now()
    if now.weekday() > 4:
        return False
    mins = now.hour * 60 + now.minute
    return 555 <= mins <= 930

# ── STARTUP ───────────────────────────────────────────────────
def show_banner(symbol, strikes, refresh):
    now = datetime.datetime.now().strftime("%d-%b-%Y %H:%M:%S")
    lines = [
        "[bold bright_cyan]NSE OPTIONS LIVE DASHBOARD v3.0[/bold bright_cyan]",
        "[dim]Powered by Dhan API v2[/dim]",
        "",
        f" Symbol  : [bold white]{symbol}[/bold white]",
        f" Strikes : [bold white]{strikes}[/bold white] above & below ATM",
        f" Refresh : [bold white]{refresh}s[/bold white]",
        f" Started : [bold white]{now}[/bold white]",
    ]
    if WATCHLIST_MODE:
        lines.append(f" Mode    : [bold yellow]WATCHLIST[/bold yellow] — {', '.join(WATCHLIST)}")
    panel = Panel(
        "\n".join(lines),
        title="[bold]Welcome[/bold]",
        border_style="bright_cyan",
        box=box.DOUBLE,
        padding=(1, 4),
    )
    console.print(panel)

def select_nearest_expiry(expiry_list):
    return sorted(expiry_list, key=lambda d: datetime.datetime.strptime(d, "%d-%b-%Y"))[0]

def test_api():
    _state.api_tested = True
    try:
        get_expiry_list(SYMBOL)
        console.print("[green]✓ Dhan API connected successfully.[/green]")
        return True
    except SystemExit:
        console.print("[red]✗ Dhan API test failed. Check credentials.[/red]")
        return False
    except Exception as e:
        console.print(f"[red]✗ Dhan API error: {e}[/red]")
        return False

# ── MAIN LOOP ─────────────────────────────────────────────────
def run():
    console.print("[cyan]Testing Dhan API connection...[/cyan]")
    if not test_api():
        raise SystemExit(1)
    time.sleep(1)
    expiry_list = get_expiry_list(SYMBOL)
    if not expiry_list:
        console.print("[red]No expiry dates found. Exiting.[/red]")
        raise SystemExit(1)
    if EXPIRY_DATE:
        nearest_expiry = EXPIRY_DATE
        console.print(f"[green]Using specified expiry: {nearest_expiry}[/green]")
    else:
        nearest_expiry = select_nearest_expiry(expiry_list)
        console.print(f"[green]Nearest expiry: {nearest_expiry}[/green]")
    time.sleep(1)
    console.clear()
    iteration = 0
    while not _state.shutdown_req:
        iteration += 1
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        _state.last_data_ts = ts
        try:
            market_open = is_market_open()
            raw = get_option_chain(SYMBOL, nearest_expiry)
            spot, rows = parse_option_chain(raw, SYMBOL)
            if spot == 0 or not rows:
                console.print("[yellow]No data received, retrying...[/yellow]")
                time.sleep(REFRESH_EVERY)
                continue
            atm_strike = get_atm_strike(spot, [r["strike"] for r in rows])
            pcr        = calc_pcr(rows)
            max_pain   = calc_max_pain(rows)
            oi_changes = get_oi_change(SYMBOL, rows)
            bias       = get_bias_label(pcr)
            bias_col   = get_bias_colour(pcr)
            check_alerts(SYMBOL, rows, pcr, oi_changes)
            title_text = (
                f"[bold white]{SYMBOL}[/bold white] | "
                f"[bold green]\u20b9{spot:.2f}[/bold green] | "
                f"Expiry: [bold cyan]{nearest_expiry}[/bold cyan] | "
                f"PCR: [bold {bias_col}]{pcr:.3f}[/bold {bias_col}] ({bias}) | "
                f"Max Pain: [bold magenta]\u20b9{max_pain:.0f}[/bold magenta] | "
                f"[bold yellow]{REFRESH_EVERY}s[/bold yellow]"
            )
            title_panel = Panel(
                title_text,
                title="[bold]NSE Options Live Dashboard[/bold]",
                border_style="bright_blue",
                box=box.DOUBLE,
                padding=(0, 2),
            )
            if not market_open:
                console.print("\n[yellow]⚠ Market Closed. Showing last data below.[/yellow]\n")
            main_table   = build_main_table(spot, rows, atm_strike, max_pain, pcr, nearest_expiry, oi_changes)
            pcr_panel    = build_pcr_panel(pcr, _state.pcr_history.get(SYMBOL, []))
            strat_panel  = build_straddle_panel(rows, atm_strike, spot)
            alerts_panel = build_alerts_panel()
            if SAVE_TO_CSV:
                save_to_csv(ts, SYMBOL, nearest_expiry, spot, rows, pcr, max_pain)
            _state.all_snapshots.append({
                "symbol": SYMBOL, "expiry": nearest_expiry,
                "spot_price": spot, "atm_strike": atm_strike,
                "max_pain": max_pain, "pcr": pcr,
                "timestamp": ts, "rows": list(rows),
            })
            if len(_state.all_snapshots) > 500:
                _state.all_snapshots.pop(0)
            console.clear()
            console.print(title_panel)
            console.print(main_table)
            console.print("\n")
            console.print(Columns([pcr_panel, strat_panel, alerts_panel], equal=False, align="top"))
            csv_count = ""
            if _state.csv_file:
                try:
                    with open(_state.csv_file) as f:
                        csv_count = f" | CSV: {sum(1 for _ in f)} rows"
                except Exception:
                    pass
            console.print(
                f"\n[dim]{SYMBOL} \u20b9{spot:.2f}"
                f" | PCR {pcr:.3f} {bias}"
                f" | Max Pain \u20b9{max_pain:.0f}"
                f" | {ts} | Next: {REFRESH_EVERY}s"
                f" | Ctrl+C to exit{csv_count}[/dim]"
            )
        except SystemExit:
            raise
        except requests.exceptions.ConnectionError:
            console.print("\n[red][ERROR] No internet connection.[/red]")
            console.print("[yellow]Retrying in 30s...[/yellow]")
            time.sleep(30)
            continue
        except Exception as e:
            console.print(f"\n[red][ERROR] {e}[/red]")
            console.print("[yellow]Retrying in 30s...[/yellow]")
            time.sleep(30)
            continue
        for _ in range(REFRESH_EVERY):
            if _state.shutdown_req:
                break
            time.sleep(1)

# ── ENTRY POINT ───────────────────────────────────────────────
def _on_ctrl_c(*_):
    _state.shutdown_req = True

if __name__ == "__main__":
    signal.signal(signal.SIGINT, _on_ctrl_c)
    console.clear()
    show_banner(SYMBOL, NUM_STRIKES, REFRESH_EVERY)
    if WATCHLIST_MODE:
        console.print("\n[yellow]Watchlist mode active.[/yellow]\n")
        run_watchlist()
    else:
        run()
    console.print("\n[cyan]Shutting down...[/cyan]")
    console.print(f"[dim]CSV file: {_state.csv_file or 'none'}[/dim]")
    console.print(f"[dim]Snapshots collected: {len(_state.all_snapshots)}[/dim]")
    if _state.all_snapshots:
        export_excel(_state.all_snapshots, SYMBOL)
    else:
        console.print("[yellow]No data to export.[/yellow]")
    console.print("[green]Goodbye![/green]")
'''

# Verify syntax
import ast
ast.parse(code)
print("[OK] SYNTAX OK - zero errors")

os.makedirs(os.path.expanduser("~/output"), exist_ok=True)
with open(os.path.expanduser("~/output/dhan_options_dashboard.py"), "w", encoding="utf-8") as f:
    f.write(code)

print(f"[INFO] {len(code.splitlines())} lines | {len(code)} chars")